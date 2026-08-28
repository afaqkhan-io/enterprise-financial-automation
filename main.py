import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from analytics_engine import run_corporate_analytics


def build_executive_dashboard(csv_source, output_excel):
    # 1. Fetch analyzed datasets from engine
    kpis, product_summary, monthly_trend = run_corporate_analytics(csv_source)

    # 2. Initialize openpyxl workbook
    wb = openpyxl.Workbook()

    # --- SHEET 1: EXECUTIVE DASHBOARD ---
    ws_dash = wb.active
    ws_dash.title = "Executive Summary"
    ws_dash.views.sheetView[0].showGridLines = True

    # Styles Definition (Enterprise Dark Navy & Slate Theme)
    brand_fill = PatternFill(
        start_color="1F4E79", end_color="1F4E79", fill_type="solid"
    )
    accent_fill = PatternFill(
        start_color="DDEBF7", end_color="DDEBF7", fill_type="solid"
    )
    zebra_fill = PatternFill(
        start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"
    )

    font_title = Font(name="Segoe UI", size=18, bold=True, color="FFFFFF")
    font_header = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    font_bold = Font(name="Segoe UI", size=11, bold=True)
    font_regular = Font(name="Segoe UI", size=11)

    thin_side = Side(border_style="thin", color="D9D9D9")
    double_side = Side(border_style="double", color="000000")
    data_border = Border(
        left=thin_side, right=thin_side, top=thin_side, bottom=thin_side
    )
    total_border = Border(top=thin_side, bottom=double_side)

    # Add Header Banner
    ws_dash.merge_cells("A1:E2")
    title_cell = ws_dash["A1"]
    title_cell.value = "ENTERPRISE PERFORMANCE DASHBOARD"
    title_cell.fill = brand_fill
    title_cell.font = font_title
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Insert KPI Cards
    ws_dash["A4"] = "KPI Metric"
    ws_dash["B4"] = "Value"
    for col in ["A4", "B4"]:
        ws_dash[col].fill = brand_fill
        ws_dash[col].font = font_header

    kpi_rows = [
        ("Gross System Revenue", kpis["Total_Gross_Revenue"], "$#,##0.00"),
        ("Discounts Conceded", kpis["Total_Discounts"], "$#,##0.00"),
        ("Net Operating Revenue", kpis["Total_Net_Revenue"], "$#,##0.00"),
        ("Net Corporate Profit", kpis["Total_Net_Profit"], "$#,##0.00"),
        ("Operational Margin", kpis["Average_Profit_Margin"] / 100, "0.00%"),
    ]

    for idx, (metric, val, fmt) in enumerate(kpi_rows, start=5):
        ws_dash[f"A{idx}"] = metric
        ws_dash[f"B{idx}"] = val
        ws_dash[f"A{idx}"].font = font_regular
        ws_dash[f"B{idx}"].font = font_bold
        ws_dash[f"B{idx}"].number_format = fmt
        ws_dash[f"A{idx}"].border = data_border
        ws_dash[f"B{idx}"].border = data_border
        if idx % 2 == 0:
            ws_dash[f"A{idx}"].fill = zebra_fill
            ws_dash[f"B{idx}"].fill = zebra_fill

    # Insert Product Breakdown Table
    ws_dash["A12"] = "Product Performance Portfolio"
    ws_dash["A12"].font = Font(name="Segoe UI", size=14, bold=True, color="1F4E79")

    headers_prod = ["Product Name", "Units Sold", "Net Revenue", "Net Profit", "Margin"]
    for c_idx, h in enumerate(headers_prod, start=1):
        cell = ws_dash.cell(row=14, column=c_idx, value=h)
        cell.fill = brand_fill
        cell.font = font_header
        cell.alignment = Alignment(horizontal="center")

    for r_idx, row in product_summary.iterrows():
        row_num = 15 + r_idx
        ws_dash.cell(
            row=row_num, column=1, value=row["Product_Name"]
        ).font = font_regular
        ws_dash.cell(
            row=row_num, column=2, value=row["Units_Sold"]
        ).number_format = "#,##0"
        ws_dash.cell(
            row=row_num, column=3, value=row["Net_Revenue"]
        ).number_format = "$#,##0.00"
        ws_dash.cell(
            row=row_num, column=4, value=row["Net_Profit"]
        ).number_format = "$#,##0.00"
        ws_dash.cell(
            row=row_num, column=5, value=row["Profit_Margin_%"] / 100
        ).number_format = "0.00%"

        for c_idx in range(1, 6):
            c = ws_dash.cell(row=row_num, column=c_idx)
            c.font = font_regular if c_idx != 1 else font_regular
            c.border = data_border
            if row_num % 2 == 0:
                c.fill = zebra_fill

    # --- SHEET 2: MONTHLY TRENDS & CHARTS ---
    ws_trend = wb.create_sheet(title="Monthly MoM Analytics")
    ws_trend.views.sheetView[0].showGridLines = True

    headers_trend = ["Month", "Net Revenue", "Net Profit", "MoM Growth %"]
    for c_idx, h in enumerate(headers_trend, start=1):
        cell = ws_trend.cell(row=1, column=c_idx, value=h)
        cell.fill = brand_fill
        cell.font = font_header
        cell.alignment = Alignment(horizontal="center")

    for r_idx, row in monthly_trend.iterrows():
        row_num = 2 + r_idx
        ws_trend.cell(row=row_num, column=1, value=row["Month"]).font = font_bold
        ws_trend.cell(
            row=row_num, column=2, value=row["Net_Revenue"]
        ).number_format = "$#,##0.00"
        ws_trend.cell(
            row=row_num, column=3, value=row["Net_Profit"]
        ).number_format = "$#,##0.00"
        ws_trend.cell(
            row=row_num, column=4, value=row["MoM_Revenue_Growth_%"] / 100
        ).number_format = "0.00%"

        for c_idx in range(1, 5):
            ws_trend.cell(row=row_num, column=c_idx).border = data_border

    # Inject Automated Corporate Bar Chart
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "Monthly Net Revenue vs Net Profit"
    chart.y_axis.title = "Financial Value ($)"
    chart.x_axis.title = "Fiscal Month"

    data_ref = Reference(
        ws_trend, min_col=2, min_row=1, max_col=3, max_row=len(monthly_trend) + 1
    )
    cats_ref = Reference(ws_trend, min_col=1, min_row=2, max_row=len(monthly_trend) + 1)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.height = 14
    chart.width = 22
    ws_trend.add_chart(chart, "F2")

    # Auto-fit column widths across sheets for clean layout
    for sheet in wb.worksheets:
        for col in sheet.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            sheet.column_dimensions[col_letter].width = max(max_len + 3, 12)

    # Save finalized professional asset
    wb.save(output_excel)
    print(
        f"\n[SYSTEM SUCCESS] Executive Automated Spreadsheet successfully written to '{output_excel}'"
    )


if __name__ == "__main__":
    build_executive_dashboard(
        "raw_sales_ledger.csv", "Corporate_Financial_Dashboard.xlsx"
    )
